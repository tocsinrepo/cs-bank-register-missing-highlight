import streamlit as st
import io
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Highlight Missing vs GL", layout="wide")
st.title("Highlight Missing Transactions vs GL")
st.write(
    "Upload a **Register** or **Bank Statement** and the **GL** (general ledger). "
    "Transactions not found in the GL will be highlighted in orange.\n\n"
    "**Supported formats:**\n"
    "- **AUB Register** – single-sheet file with Date, Page, Description, "
    "Debits (Out), Credits (In) columns\n"
    "- **AUB Bank Statement** – multi-sheet file with `purch1` (withdrawals) "
    "and `dep1` (deposits) sheets\n\n"
    "Checks are matched by **check number**. Other transactions are matched "
    "by **amount and date** (within a 5-day window). "
    "Transfers ending in **7459** are ignored."
)

ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
TRANSFER_7459_KEYWORD = "7459"
DATE_TOLERANCE_DAYS = 5

# --------------- Format detection ---------------

FORMAT_REGISTER = "register"
FORMAT_STATEMENT = "statement"


def detect_format(wb):
    """Auto-detect whether the workbook is a Register or a Bank Statement.

    Bank Statement files have sheets named 'purch1' and 'dep1'.
    Register files have a single sheet with header row containing 'Debits (Out)'.
    """
    sheet_names = [s.lower() for s in wb.sheetnames]
    if "purch1" in sheet_names and "dep1" in sheet_names:
        return FORMAT_STATEMENT
    return FORMAT_REGISTER


# --------------- Shared helpers ---------------


def parse_date(value):
    """Parse a date value from either a datetime object or a string like '3/18/2026'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def extract_gl_data(wb):
    """Extract transaction data from the GL file.

    GL layout (columns by index):
      I (9)  = Type, K (11) = Date, M (13) = Num,
      U (21) = Debit, W (23) = Credit
    In accounting for a cash/bank account:
      GL Debit  = money coming IN  (matches register credits / statement deposits)
      GL Credit = money going OUT  (matches register debits / statement withdrawals)

    Returns:
      gl_debits:  list of (date, amount) for non-check deposits/credits in
      gl_credits: list of (date, amount) for non-check payments out
      gl_checks:  list of (check_number_str, amount) for check payments
    """
    ws = wb.active
    gl_debits = []
    gl_credits = []
    gl_checks = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_val = row[10].value if len(row) > 10 else None
        num_val = row[12].value if len(row) > 12 else None
        debit_val = row[20].value if len(row) > 20 else None
        credit_val = row[22].value if len(row) > 22 else None

        if date_val is None or date_val == "Date":
            continue

        dt = parse_date(date_val)
        if dt is None:
            continue

        num_str = str(num_val).strip() if num_val is not None else ""
        is_gl_check = bool(re.match(r"^\d{4,6}$", num_str))

        if credit_val is not None:
            try:
                amt = round(float(credit_val), 2)
                if is_gl_check:
                    gl_checks.append((num_str, amt))
                else:
                    gl_credits.append((dt, amt))
            except (ValueError, TypeError):
                pass

        if debit_val is not None:
            try:
                gl_debits.append((dt, round(float(debit_val), 2)))
            except (ValueError, TypeError):
                pass

    return gl_debits, gl_credits, gl_checks


def extract_check_number(description):
    """Extract the check number from a description like 'CHECK - 10442' or 'CHECK #10442'."""
    if description is None:
        return None
    desc = str(description)
    if "CHECK" not in desc.upper():
        return None
    m = re.search(r"(\d{4,6})", desc)
    return m.group(1) if m else None


def is_transfer_7459(description):
    """Check if a transaction description refers to a transfer ending in 7459."""
    if description is None:
        return False
    return TRANSFER_7459_KEYWORD in str(description)


def find_match(reg_date, reg_amount, pool, tolerance_days=DATE_TOLERANCE_DAYS):
    """Find and remove a matching (date, amount) entry from the pool.

    A match requires the same amount and a date within tolerance_days.
    If multiple matches exist, prefer the closest date.
    Returns True if a match was found and consumed.
    """
    best_idx = None
    best_diff = None

    for i, (gl_date, gl_amount) in enumerate(pool):
        if gl_amount == reg_amount:
            if reg_date is not None and gl_date is not None:
                diff = abs((reg_date - gl_date).days)
                if diff <= tolerance_days:
                    if best_diff is None or diff < best_diff:
                        best_idx = i
                        best_diff = diff
            elif reg_date is None or gl_date is None:
                if best_idx is None:
                    best_idx = i
                    best_diff = 0

    if best_idx is not None:
        pool.pop(best_idx)
        return True
    return False


def find_check_match(check_num, pool):
    """Find and remove a matching check number entry from the GL check pool."""
    for i, (gl_check_num, gl_amount) in enumerate(pool):
        if gl_check_num == check_num:
            pool.pop(i)
            return True
    return False


# --------------- Register format processing ---------------


def process_register(register_wb, gl_debits, gl_credits, gl_checks):
    """Process a single-sheet AUB Register file.

    Register layout:
      Col 1 = Date, Col 2 = Page, Col 3 = Description,
      Col 4 = Debits (Out), Col 5 = Credits (In), Col 6 = Balance

    Returns the modified workbook and counts.
    """
    ws = register_wb.active

    gl_credit_pool = list(gl_credits)
    gl_debit_pool = list(gl_debits)
    gl_check_pool = list(gl_checks)

    total = 0
    missing = 0
    ignored = 0

    for row_idx in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=1).value
        if date_cell is None:
            continue
        date_str = str(date_cell)
        if any(kw in date_str for kw in ["TOTALS", "Total", "Beginning", "Ending", "balance"]):
            continue

        reg_date = parse_date(date_cell)
        description = ws.cell(row=row_idx, column=3).value or ""
        debit_val = ws.cell(row=row_idx, column=4).value
        credit_val = ws.cell(row=row_idx, column=5).value

        if is_transfer_7459(description):
            ignored += 1
            continue

        total += 1
        is_missing = False
        check_num = extract_check_number(description)

        if debit_val is not None:
            try:
                amt = round(float(debit_val), 2)
                if amt != 0:
                    if check_num:
                        if not find_check_match(check_num, gl_check_pool):
                            is_missing = True
                    else:
                        if not find_match(reg_date, amt, gl_credit_pool):
                            is_missing = True
            except (ValueError, TypeError):
                pass

        if credit_val is not None:
            try:
                amt = round(float(credit_val), 2)
                if amt != 0:
                    if not find_match(reg_date, amt, gl_debit_pool):
                        is_missing = True
            except (ValueError, TypeError):
                pass

        if is_missing:
            missing += 1
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ORANGE_FILL

    return register_wb, total, missing, ignored


# --------------- Statement format processing ---------------


def _skip_row_statement(date_val, desc):
    """Return True if this row should be skipped (headers, totals, blanks)."""
    if date_val is None:
        return True
    s = str(date_val)
    if any(kw in s for kw in ["TOTAL", "GRAND", "CHECKS", "SWEEPS", "ACH", "REGULAR"]):
        return True
    d = str(desc) if desc else ""
    if any(kw in d for kw in ["TOTAL", "GRAND"]):
        return True
    return False


def process_statement(stmt_wb, gl_debits, gl_credits, gl_checks):
    """Process a multi-sheet AUB Bank Statement file.

    Reads the dep1 sheet (deposits) and purch1 sheet (withdrawals) separately.
    Highlights missing rows orange on each sheet.

    dep1 layout:
      Col 1 = Date, Col 2 = Description, Col 4 = Amount
      All amounts are deposits (money in) → match against GL debits.

    purch1 layout:
      Col 2 = Date, Col 3 = Description, Col 4 = Amount
      All amounts are withdrawals (money out) → match against GL credits or checks.

    Returns the modified workbook and combined counts.
    """
    gl_debit_pool = list(gl_debits)
    gl_credit_pool = list(gl_credits)
    gl_check_pool = list(gl_checks)

    total = 0
    missing = 0
    ignored = 0
    dep_missing = 0
    purch_missing = 0

    # ---- dep1 (Deposits) ----
    dep_ws = stmt_wb["dep1"]
    for row_idx in range(3, dep_ws.max_row + 1):
        date_val = dep_ws.cell(row=row_idx, column=1).value
        desc = dep_ws.cell(row=row_idx, column=2).value or ""
        amt_val = dep_ws.cell(row=row_idx, column=4).value

        if _skip_row_statement(date_val, desc):
            continue
        if amt_val is None:
            continue

        reg_date = parse_date(date_val)
        try:
            amt = round(float(amt_val), 2)
        except (ValueError, TypeError):
            continue

        if is_transfer_7459(desc):
            ignored += 1
            continue

        total += 1

        if not find_match(reg_date, amt, gl_debit_pool):
            missing += 1
            dep_missing += 1
            for c in range(1, dep_ws.max_column + 1):
                dep_ws.cell(row=row_idx, column=c).fill = ORANGE_FILL

    # ---- purch1 (Withdrawals) ----
    purch_ws = stmt_wb["purch1"]
    for row_idx in range(3, purch_ws.max_row + 1):
        date_val = purch_ws.cell(row=row_idx, column=2).value
        desc = purch_ws.cell(row=row_idx, column=3).value or ""
        amt_val = purch_ws.cell(row=row_idx, column=4).value

        if _skip_row_statement(date_val, desc):
            continue
        if amt_val is None:
            continue

        reg_date = parse_date(date_val)
        try:
            amt = round(float(amt_val), 2)
        except (ValueError, TypeError):
            continue

        if is_transfer_7459(desc):
            ignored += 1
            continue

        total += 1
        check_num = extract_check_number(desc)
        is_missing = False

        if check_num:
            if not find_check_match(check_num, gl_check_pool):
                is_missing = True
        else:
            if not find_match(reg_date, amt, gl_credit_pool):
                is_missing = True

        if is_missing:
            missing += 1
            purch_missing += 1
            for c in range(1, purch_ws.max_column + 1):
                purch_ws.cell(row=row_idx, column=c).fill = ORANGE_FILL

    return stmt_wb, total, missing, ignored, dep_missing, purch_missing


# --------------- Streamlit UI ---------------

col1, col2 = st.columns(2)
with col1:
    register_file = st.file_uploader(
        "Upload Register or Bank Statement (Excel)", type=["xlsx", "xls"]
    )
with col2:
    gl_file = st.file_uploader("Upload GL (Excel)", type=["xlsx", "xls"])

if register_file and gl_file:
    if st.button("Compare & Highlight Missing", type="primary"):
        with st.spinner("Loading GL..."):
            gl_wb = load_workbook(io.BytesIO(gl_file.read()), data_only=True)
            gl_debits, gl_credits, gl_checks = extract_gl_data(gl_wb)

        st.info(
            f"GL loaded: **{len(gl_debits)}** deposit/debit entries, "
            f"**{len(gl_credits)}** non-check credit entries, "
            f"**{len(gl_checks)}** check entries."
        )

        with st.spinner("Detecting format and comparing..."):
            file_bytes = io.BytesIO(register_file.read())
            wb = load_workbook(file_bytes)
            fmt = detect_format(wb)

        if fmt == FORMAT_REGISTER:
            st.info("Detected format: **AUB Register** (single-sheet)")
            with st.spinner("Comparing register against GL..."):
                wb, total, missing, ignored = process_register(
                    wb, gl_debits, gl_credits, gl_checks
                )
            matched = total - missing
            st.success(
                f"Done! **{total}** transactions checked, **{matched}** matched, "
                f"**{missing}** missing (highlighted orange), "
                f"**{ignored}** transfers to 7459 ignored."
            )

        else:
            st.info("Detected format: **AUB Bank Statement** (multi-sheet with dep1 & purch1)")
            with st.spinner("Comparing statement against GL..."):
                wb, total, missing, ignored, dep_miss, purch_miss = process_statement(
                    wb, gl_debits, gl_credits, gl_checks
                )
            matched = total - missing
            st.success(
                f"Done! **{total}** transactions checked, **{matched}** matched, "
                f"**{missing}** missing (highlighted orange), "
                f"**{ignored}** transfers to 7459 ignored."
            )
            if missing > 0:
                st.write(
                    f"- **{dep_miss}** missing deposit(s) on `dep1` sheet\n"
                    f"- **{purch_miss}** missing withdrawal(s) on `purch1` sheet"
                )

        if missing > 0:
            st.warning(
                f"{missing} transaction(s) highlighted in orange are missing from the GL."
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        base_name = register_file.name.rsplit(".", 1)[0] if register_file.name else "file"
        download_name = f"{base_name}_highlighted.xlsx"

        st.download_button(
            label="Download Highlighted File",
            data=output.getvalue(),
            file_name=download_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheetml",
            type="primary",
        )
